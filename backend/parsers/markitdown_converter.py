"""Markitdown converter module for DOCX/DOC/PDF/XLSX to Markdown conversion."""

import base64
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# markdownify treats images inside table cells / headings as inline decorations
# and by default (keep_inline_images_in=[]) drops them, keeping only the alt
# text.  Bid documents routinely wrap certificate scans and verification
# screenshots in (borderless) tables; losing those refs made the review agent
# see empty sections and report missing evidence.  Whitelist the tags mammoth
# uses around cell images so `![alt](path)` refs survive inside tables.
_KEEP_INLINE_IMAGES_IN = ["p", "td", "th", "h1", "h2", "h3", "h4", "h5", "h6"]

# 1x1 transparent PNG substituted for DOCX images whose part is missing or
# unreadable inside the archive (e.g. bid-authoring tools emitting image
# relationships with Target="../NULL").
_PLACEHOLDER_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJ"
    "AAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
)

# XLSX 输出上限：标书里的报价清单表动辄上千行，而审查流程对解析产物无
# 分段截断（全文进 LLM），全量转换会使 token 失控并易触发解析软超时。
# 超出部分在 Markdown 中以省略提示标注，保留可追溯性。可通过
# MarkitdownConverter 构造参数（来自 Settings，环境变量可覆盖）调整。
XLSX_MAX_ROWS_PER_SHEET = 300
XLSX_MAX_COLS_PER_SHEET = 64


def _xlsx_is_empty_row(row_values) -> bool:
    """A row with only None/blank cells: skipped entirely in sheet output."""
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row_values
    )


def _xlsx_cell_text(value) -> str:
    """Render one cell value as Markdown-table-safe text.

    ``data_only=True`` makes formula cells carry their Excel-cached result;
    uncached formulas and merged-cell non-anchor cells read back as None.
    """
    import datetime as dt
    import math

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float):
        # 金额/数量列以整数为主：1.0 显示为 1 更贴近 Excel 展示，也省 token
        if math.isfinite(value) and value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return str(value)
    text = str(value)
    # 竖线会破坏 Markdown 表格结构；单元格内换行替换为 <br> 保留多行语义
    text = text.replace("|", "\\|")
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    return text


class MarkitdownConversionError(Exception):
    """Raised when markitdown conversion fails."""
    pass


@dataclass
class ImageInfo:
    """Image information extracted from document."""
    filename: str
    data: bytes


@dataclass
class ConversionResult:
    """Result of document conversion."""
    markdown_content: str
    images: list[ImageInfo]
    page_count: Optional[int] = None


class DirectFileImageHandler:
    """Writes DOCX images directly to disk, bypassing base64 encoding.

    Instead of letting mammoth base64-encode images into data URIs, this handler
    reads image bytes from the DOCX and writes them directly to the target directory,
    returning lightweight file-path references for the HTML/Markdown output.
    """

    def __init__(self, images_dir: Path, images_dir_name: str):
        self._images_dir = images_dir
        self._images_dir_name = images_dir_name
        self._counter = 0
        self.images: list[ImageInfo] = []
        self.failed_images = 0

    def __call__(self, image):
        self._counter += 1
        content_type = str(image.content_type or "").lower()
        ext = {
            "image/jpeg": "jpeg",
            "image/jpg": "jpg",
            "image/png": "png",
            "image/gif": "gif",
            "image/webp": "webp",
            "image/bmp": "bmp",
            "image/tiff": "tiff",
            "image/svg+xml": "svg",
            "image/jp2": "jp2",
            "image/jpx": "jpx",
        }.get(content_type, content_type.partition("/")[2] or "png")
        ext = ext.split("+", 1)[0]
        filename = f"image_{self._counter}.{ext}"

        self._images_dir.mkdir(parents=True, exist_ok=True)
        try:
            with image.open() as image_bytes:
                data = image_bytes.read()
        except Exception as e:
            # Some authoring tools emit image relationships pointing at parts
            # that don't exist in the archive (Target="../NULL"); mammoth
            # surfaces that as KeyError from zipfile.  One broken image must
            # not fail the whole conversion: substitute a 1x1 placeholder so
            # markdown refs, files on disk and the images list stay consistent.
            logger.warning(
                "DOCX image %s could not be read (%s); substituting 1x1 placeholder",
                filename,
                e,
            )
            self.failed_images += 1
            filename = f"image_{self._counter}.png"
            data = _PLACEHOLDER_PNG

        dest = self._images_dir / filename
        dest.write_bytes(data)

        self.images.append(ImageInfo(filename=filename, data=data))
        return {"src": f"{self._images_dir_name}/{filename}"}


class MarkitdownConverter:
    """Markitdown converter for DOCX/DOC/PDF/XLSX files to Markdown format.

    Uses the markitdown library to extract text and images from documents.
    For DOCX/DOC, images are written directly to disk via a custom mammoth image handler.
    For PDF, uses markitdown's built-in PdfConverter for text/tables and PyMuPDF for images.
    For XLSX, converts each visible worksheet to a Markdown table via openpyxl
    (cached formula values; row/column caps; embedded images materialized like DOCX).
    """

    def __init__(
        self,
        timeout: int = 300,
        xlsx_max_rows: Optional[int] = None,
        xlsx_max_cols: Optional[int] = None,
    ):
        self.timeout = timeout
        self.xlsx_max_rows = xlsx_max_rows if xlsx_max_rows is not None else XLSX_MAX_ROWS_PER_SHEET
        self.xlsx_max_cols = xlsx_max_cols if xlsx_max_cols is not None else XLSX_MAX_COLS_PER_SHEET

    def convert(self, file_path: Path, progress_callback=None, images_dir: Path = None) -> ConversionResult:
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在，请重新上传：{file_path}")

        suffix = file_path.suffix.lower()
        if suffix not in [".docx", ".doc", ".pdf", ".xlsx"]:
            raise ValueError(f"暂不支持 {suffix or '未知'} 格式，请上传 DOCX、DOC、PDF 或 XLSX 文件")

        file_size = file_path.stat().st_size
        logger.info(f"Markitdown conversion: {file_path} ({file_size / (1024 * 1024):.2f}MB)")

        if images_dir is None:
            images_dir = file_path.parent / f"{file_path.stem}_images"

        try:
            if suffix in [".docx", ".doc"]:
                return self._convert_docx(file_path, images_dir, progress_callback)
            elif suffix == ".xlsx":
                return self._convert_xlsx(file_path, images_dir, progress_callback)
            else:
                return self._convert_pdf(file_path, images_dir, progress_callback)

        except MarkitdownConversionError:
            raise
        except Exception as e:
            logger.error(f"Markitdown conversion failed: {e}")
            raise MarkitdownConversionError(f"文档转换失败：{e}")

    def _convert_docx(self, file_path: Path, images_dir: Path, progress_callback) -> ConversionResult:
        """Convert DOCX/DOC while materializing every embedded image.

        MarkItDown's DOCX converter accepts arbitrary keyword arguments but its
        current implementation does not forward ``convert_image`` to Mammoth.
        Calling it therefore produces ``data:image/...;base64`` Markdown and
        leaves the image directory empty.  Invoke Mammoth directly, then apply
        the same HTML-to-Markdown stage so image references remain lightweight
        paths backed by real files.
        """
        import mammoth
        from markdownify import markdownify

        images_dir_name = images_dir.name
        handler = DirectFileImageHandler(images_dir, images_dir_name)
        with file_path.open("rb") as source:
            html_result = mammoth.convert_to_html(
                source,
                convert_image=mammoth.images.img_element(handler),
            )
        for message in html_result.messages:
            logger.warning("Mammoth DOCX conversion warning: %s", message.message)
        markdown_content = markdownify(
            html_result.value,
            heading_style="ATX",
            bullets="-",
            keep_inline_images_in=_KEEP_INLINE_IMAGES_IN,
        ).strip()
        if progress_callback:
            progress_callback(1, 1)

        logger.info(
            "DOCX conversion successful: %s chars, %s materialized images",
            len(markdown_content),
            len(handler.images),
        )
        if handler.failed_images:
            logger.warning(
                "DOCX contained %s unreadable image(s); substituted 1x1 placeholder(s)",
                handler.failed_images,
            )

        return ConversionResult(
            markdown_content=markdown_content,
            images=handler.images,
            page_count=None,
        )

    def _convert_xlsx(self, file_path: Path, images_dir: Path, progress_callback) -> ConversionResult:
        """Convert XLSX to Markdown: one section per visible worksheet.

        - data_only=True reads cached formula results; a formula whose value was
          never cached by Excel reads back as an empty cell.
        - Output per sheet is capped at xlsx_max_rows data rows / xlsx_max_cols
          columns; omitted ranges are flagged inline so the review agent knows
          coverage is partial (review consumes the full markdown, unchunked).
        - Embedded images are materialized to images_dir under the same contract
          as DOCX — certificate screenshots inside spreadsheets are review
          evidence (see the 2026-08-18 missing-table-images incident).  Chart
          objects cannot be rendered without Excel and are only flagged.
        """
        import zipfile

        from openpyxl import load_workbook

        try:
            workbook = load_workbook(filename=str(file_path), data_only=True)
        except zipfile.BadZipFile as exc:
            # 密码保护的 xlsx 是 OLE2 加密容器而非 zip，openpyxl 同样报 BadZipFile
            raise MarkitdownConversionError(
                "Excel 文件无法读取（文件已损坏或为加密文件），请解除密码后另存为 .xlsx 重新上传"
            ) from exc

        images: list[ImageInfo] = []
        sections: list[str] = []
        hidden_sheets: list[str] = []
        sheet_names = list(workbook.sheetnames)

        for sheet_index, sheet_name in enumerate(sheet_names, start=1):
            worksheet = workbook[sheet_name]
            if worksheet.sheet_state != "visible":
                hidden_sheets.append(str(sheet_name))
                continue
            if progress_callback:
                progress_callback(sheet_index, len(sheet_names))
            sections.append(
                self._xlsx_sheet_to_markdown(worksheet, sheet_index, images_dir, images)
            )

        workbook.close()

        markdown_parts = sections or ["（Excel 文件中没有可见工作表）"]
        if hidden_sheets:
            markdown_parts.append(
                f"> 注：已跳过 {len(hidden_sheets)} 个隐藏工作表（{'、'.join(hidden_sheets)}）"
            )
        markdown_content = "\n\n".join(markdown_parts)

        if progress_callback:
            progress_callback(len(sheet_names), len(sheet_names))

        logger.info(
            "XLSX conversion successful: %s chars, %s materialized images, %s visible sheets",
            len(markdown_content),
            len(images),
            len(sections),
        )

        return ConversionResult(
            markdown_content=markdown_content,
            images=images,
            page_count=None,
        )

    def _xlsx_sheet_to_markdown(
        self,
        worksheet,
        sheet_index: int,
        images_dir: Path,
        images: list[ImageInfo],
    ) -> str:
        from openpyxl.utils import get_column_letter

        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        effective_cols = min(max_col, self.xlsx_max_cols)
        col_truncated = max_col > self.xlsx_max_cols

        table_lines: list[str] = []
        if effective_cols > 0:
            table_lines.append(
                "| 行号 | "
                + " | ".join(
                    get_column_letter(col) for col in range(1, effective_cols + 1)
                )
                + " |"
            )
            table_lines.append(
                "| " + " | ".join(["---"] * (effective_cols + 1)) + " |"
            )

        kept_rows = 0
        rows_truncated = False
        if effective_cols > 0:
            row_iterator = worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                max_col=effective_cols,
                values_only=True,
            )
            for row_index, row_values in enumerate(row_iterator, start=1):
                if _xlsx_is_empty_row(row_values):
                    # 空行不占用行数上限也不输出：格式化撑大的已用区域里空行居多
                    continue
                if kept_rows >= self.xlsx_max_rows:
                    rows_truncated = True
                    break
                kept_rows += 1
                table_lines.append(
                    "| "
                    + " | ".join([str(row_index)] + [_xlsx_cell_text(v) for v in row_values])
                    + " |"
                )

        notes: list[str] = []
        if col_truncated:
            notes.append(
                f"> ⚠️ 该工作表共 {max_col} 列，超过列数上限 {self.xlsx_max_cols}，右侧列已省略"
            )
        if rows_truncated:
            notes.append(
                f"> ⚠️ 该工作表数据行数超过行数上限 {self.xlsx_max_rows}，"
                f"仅保留前 {kept_rows} 个非空数据行，其余已省略"
            )
        chart_count = len(getattr(worksheet, "_charts", []) or [])
        if chart_count:
            notes.append(
                f"> ⚠️ 该工作表含 {chart_count} 个图表对象，当前版本未提取图表图像"
            )

        image_refs = self._extract_xlsx_images(worksheet, sheet_index, images_dir, images)

        parts = [f"## {worksheet.title}"]
        if kept_rows:
            # 表格行必须连续（空行会终止 GFM 表格），整表作为一个块参与分节 join
            parts.append("\n".join(table_lines))
        else:
            parts.append("（空工作表）")
        parts.extend(notes)
        parts.extend(image_refs)
        return "\n\n".join(parts)

    @staticmethod
    def _extract_xlsx_images(
        worksheet,
        sheet_index: int,
        images_dir: Path,
        images: list[ImageInfo],
    ) -> list[str]:
        """Materialize embedded worksheet images to disk (DOCX-equivalent contract).

        openpyxl only populates ws._images when Pillow is available; exotic
        formats are re-encoded by Image._data().  One broken image must not
        fail the sheet conversion.
        """
        image_refs: list[str] = []
        for image_index, drawing_image in enumerate(
            getattr(worksheet, "_images", []) or [], start=1
        ):
            try:
                data = drawing_image._data()
                image_format = str(getattr(drawing_image, "format", None) or "png").lower()
                ext = {"jpeg": "jpg", "jpg": "jpg"}.get(image_format, image_format)
                filename = f"sheet{sheet_index}_img{image_index}.{ext}"
                images_dir.mkdir(parents=True, exist_ok=True)
                (images_dir / filename).write_bytes(data)
                images.append(ImageInfo(filename=filename, data=data))
                row_hint = ""
                anchor_from = getattr(getattr(drawing_image, "anchor", None), "_from", None)
                if anchor_from is not None and getattr(anchor_from, "row", None) is not None:
                    row_hint = f"（约第 {anchor_from.row + 1} 行）"
                image_refs.append(
                    f"![工作表内嵌图片 sheet{sheet_index}_img{image_index} {row_hint}]"
                    f"({images_dir.name}/{filename})"
                )
            except Exception as exc:
                logger.warning(
                    "XLSX image %s of sheet %s could not be read: %s",
                    image_index,
                    sheet_index,
                    exc,
                )
        return image_refs

    def _convert_pdf(self, file_path: Path, images_dir: Path, progress_callback=None) -> ConversionResult:
        """Convert PDF file to Markdown using PyMuPDF page-by-page extraction.

        Extracts text and images per page, inserting image references inline
        after each page's text content.
        """
        import fitz

        images_dir.mkdir(parents=True, exist_ok=True)
        images_dir_name = images_dir.name

        page_parts: list[str] = []
        all_images: list[ImageInfo] = []

        doc = fitz.open(str(file_path))
        if doc.is_encrypted:
            doc.close()
            raise MarkitdownConversionError("PDF 文件已加密，请解除密码后重新上传")

        total_pages = len(doc)
        logger.info(f"PDF fitz conversion: {file_path.name}, {total_pages} pages")

        for page_num in range(total_pages):
            page = doc[page_num]
            page_text = page.get_text().strip()
            page_image_refs: list[str] = []

            try:
                for img_index, img in enumerate(page.get_images()):
                    try:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        image_ext = base_image["ext"]
                        if len(image_bytes) > 10 * 1024 * 1024:
                            continue
                        filename = f"page_{page_num + 1}_img_{img_index + 1}.{image_ext}"
                        (images_dir / filename).write_bytes(image_bytes)
                        all_images.append(ImageInfo(filename=filename, data=image_bytes))
                        page_image_refs.append(f"![{filename}]({images_dir_name}/{filename})")
                    except Exception as e:
                        logger.warning(f"Failed to extract image {img_index} from page {page_num + 1}: {e}")
            except Exception as e:
                logger.warning(f"Failed to get images from page {page_num + 1}: {e}")

            if page_text:
                page_parts.append(page_text)
            page_parts.extend(page_image_refs)

            if progress_callback:
                progress_callback(page_num + 1, total_pages)

        doc.close()

        markdown_content = "\n\n".join(page_parts)
        logger.info(f"PDF fitz conversion result: {len(markdown_content)} chars, {len(all_images)} images, {total_pages} pages")

        return ConversionResult(
            markdown_content=markdown_content,
            images=all_images,
            page_count=total_pages,
        )


def convert_to_markdown(file_path: Path) -> ConversionResult:
    converter = MarkitdownConverter()
    return converter.convert(file_path)


def _extract_pdf_text_with_fitz(file_path: Path) -> str:
    """Extract text from PDF using PyMuPDF as a fallback when markitdown fails."""
    import fitz

    text_parts = []
    try:
        doc = fitz.open(str(file_path))
        if doc.is_encrypted:
            doc.close()
            raise ValueError("PDF 文件已加密，请解除密码后重新上传")

        max_pages = 500
        for page_num in range(min(len(doc), max_pages)):
            page = doc[page_num]
            try:
                text = page.get_text()
                if text and text.strip():
                    text_parts.append(text.strip())
            except Exception as e:
                logger.warning(f"Failed to extract text from page {page_num + 1}: {e}")
        doc.close()
    except Exception as e:
        logger.error(f"PyMuPDF text extraction failed for {file_path}: {e}")
        raise

    return "\n\n".join(text_parts)


def _extract_pdf_images(file_path: Path, images_dir: Path) -> list[ImageInfo]:
    """Extract embedded images from PDF using PyMuPDF."""
    import fitz

    images: list[ImageInfo] = []

    try:
        doc = fitz.open(str(file_path))
        if doc.is_encrypted:
            logger.warning(f"PDF is encrypted, skipping image extraction: {file_path}")
            doc.close()
            return images

        max_pages = 500
        for page_num in range(min(len(doc), max_pages)):
            page = doc[page_num]
            try:
                for img_index, img in enumerate(page.get_images()):
                    try:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        image_ext = base_image["ext"]
                        if len(image_bytes) > 10 * 1024 * 1024:
                            continue
                        filename = f"page_{page_num + 1}_img_{img_index + 1}.{image_ext}"
                        images_dir.mkdir(parents=True, exist_ok=True)
                        (images_dir / filename).write_bytes(image_bytes)
                        images.append(ImageInfo(filename=filename, data=image_bytes))
                    except Exception as e:
                        logger.warning(f"Failed to extract image {img_index} from page {page_num + 1}: {e}")
            except Exception as e:
                logger.warning(f"Failed to get images from page {page_num + 1}: {e}")

        doc.close()
    except Exception as e:
        logger.warning(f"PDF image extraction failed for {file_path}: {e}")

    return images


def _get_pdf_page_count(file_path: Path) -> Optional[int]:
    """Get PDF page count using PyMuPDF."""
    import fitz

    try:
        doc = fitz.open(str(file_path))
        count = len(doc)
        doc.close()
        return count
    except Exception as e:
        logger.warning(f"Failed to get PDF page count: {e}")
        return None
