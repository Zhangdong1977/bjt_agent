"""Tests for MarkitdownConverter XLSX support (worksheets -> markdown tables)."""

import datetime as dt
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.parsers.markitdown_converter import (
    XLSX_MAX_COLS_PER_SHEET,
    XLSX_MAX_ROWS_PER_SHEET,
    MarkitdownConversionError,
    MarkitdownConverter,
)


def _write_workbook(path: Path, populate) -> Path:
    """Create an xlsx file at *path*, delegating cell writing to *populate*."""
    workbook = Workbook()
    populate(workbook)
    workbook.save(path)
    return path


class TestConvertXlsx:

    def test_multi_sheet_values_and_structure(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            sheet.title = "报价表"
            sheet["A1"] = "序号"
            sheet["B1"] = "项目名称"
            sheet["C1"] = "金额"
            sheet["A2"] = 1
            sheet["B2"] = "设备采购"
            sheet["C2"] = 123000.0
            sheet["A3"] = 2
            sheet["B3"] = "安装调试"
            sheet["C3"] = 4500.25
            other = workbook.create_sheet("第二页")
            other["A1"] = "备注"
            other["B1"] = True
            other["A2"] = dt.datetime(2026, 8, 22, 14, 30, 0)
            other["B2"] = dt.date(2026, 8, 23)

        file_path = _write_workbook(tmp_path / "bid.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        assert "## 报价表" in result.markdown_content
        assert "## 第二页" in result.markdown_content
        # 表头带行号列与 Excel 列字母，便于审查引用定位
        assert "| 行号 | A | B | C |" in result.markdown_content
        assert "| 1 | 序号 | 项目名称 | 金额 |" in result.markdown_content
        # 表格行必须连续：空行会终止 GFM 表格，导致逐行渲染成孤立表格
        assert "|\n| 2 | 1 | 设备采购 | 123000 |" in result.markdown_content
        # 整数金额显示为整数（123000.0 -> 123000），小数保留
        assert "| 2 | 1 | 设备采购 | 123000 |" in result.markdown_content
        assert "4500.25" in result.markdown_content
        # 布尔与日期时间格式
        assert "TRUE" in result.markdown_content
        assert "2026-08-22 14:30:00" in result.markdown_content
        assert "2026-08-23" in result.markdown_content
        # xlsx 无页概念，与 DOCX 契约一致
        assert result.page_count is None
        assert result.images == []

    def test_merged_cells_anchor_value_only(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "合并标题"
            sheet.merge_cells("A1:C1")
            sheet["A2"] = "x"
            sheet["C2"] = "y"

        file_path = _write_workbook(tmp_path / "merged.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        # 合并区域只有左上角携带值，非锚点单元格读取为 None -> 空单元格
        assert "| 1 | 合并标题 |  |  |" in result.markdown_content
        assert "| 2 | x |  | y |" in result.markdown_content

    def test_formula_without_cached_value_renders_empty(self, tmp_path):
        # openpyxl 写入的公式没有 Excel 缓存值，data_only=True 读回 None。
        # 真实用户文件通常由 Excel 保存过，公式单元格会带缓存结果。
        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "合计"
            sheet["B1"] = "=A2+A3"
            sheet["A2"] = 1
            sheet["A3"] = 2

        file_path = _write_workbook(tmp_path / "formula.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        assert "合计" in result.markdown_content
        # B1 无缓存 -> 空单元格，而非公式文本
        assert "=A2+A3" not in result.markdown_content

    def test_row_cap_with_omission_note(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            for row in range(1, 11):
                sheet.cell(row=row, column=1, value=f"行{row}")

        file_path = _write_workbook(tmp_path / "rows.xlsx", populate)
        result = MarkitdownConverter(xlsx_max_rows=3).convert(file_path)

        assert "| 1 | 行1 |" in result.markdown_content
        assert "| 3 | 行3 |" in result.markdown_content
        assert "行4" not in result.markdown_content
        assert "数据行数超过行数上限 3" in result.markdown_content
        assert "仅保留前 3 个非空数据行" in result.markdown_content

    def test_row_cap_exact_rows_no_note(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            for row in range(1, 4):
                sheet.cell(row=row, column=1, value=f"行{row}")

        file_path = _write_workbook(tmp_path / "exact.xlsx", populate)
        result = MarkitdownConverter(xlsx_max_rows=3).convert(file_path)

        assert "已省略" not in result.markdown_content
        assert "行数上限" not in result.markdown_content

    def test_empty_rows_do_not_consume_cap(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "标题"
            sheet["A5"] = "第二行数据"

        file_path = _write_workbook(tmp_path / "sparse.xlsx", populate)
        result = MarkitdownConverter(xlsx_max_rows=2).convert(file_path)

        # 中间的空行（2/3/4 行）被跳过且不计入上限；行号保留原始 Excel 行号
        assert "| 1 | 标题 |" in result.markdown_content
        assert "| 5 | 第二行数据 |" in result.markdown_content
        assert "已省略" not in result.markdown_content

    def test_column_cap_with_omission_note(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            for col in range(1, 11):
                sheet.cell(row=1, column=col, value=f"列{col}")

        file_path = _write_workbook(tmp_path / "cols.xlsx", populate)
        result = MarkitdownConverter(xlsx_max_cols=5).convert(file_path)

        assert "列5" in result.markdown_content
        assert "列6" not in result.markdown_content
        assert f"该工作表共 10 列，超过列数上限 5" in result.markdown_content

    def test_narrow_sheet_table_width_follows_used_columns(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "a"
            sheet["B1"] = "b"

        file_path = _write_workbook(tmp_path / "narrow.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        # 列宽跟随实际已用列数，而不是上限（省 token）
        assert "| 行号 | A | B |" in result.markdown_content
        assert "| C |" not in result.markdown_content

    def test_empty_sheet_placeholder(self, tmp_path):
        def populate(workbook):
            workbook.active.title = "空白页"
            workbook.create_sheet("有数据")
            workbook["有数据"]["A1"] = "内容"

        file_path = _write_workbook(tmp_path / "empty.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        assert "## 空白页" in result.markdown_content
        assert "（空工作表）" in result.markdown_content
        assert "## 有数据" in result.markdown_content

    def test_hidden_sheet_skipped_with_note(self, tmp_path):
        def populate(workbook):
            workbook.active.title = "可见页"
            workbook["可见页"]["A1"] = "内容"
            hidden = workbook.create_sheet("隐藏页")
            hidden.sheet_state = "hidden"
            hidden["A1"] = "不应出现"

        file_path = _write_workbook(tmp_path / "hidden.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        assert "## 可见页" in result.markdown_content
        assert "## 隐藏页" not in result.markdown_content
        assert "不应出现" not in result.markdown_content
        assert "已跳过 1 个隐藏工作表（隐藏页）" in result.markdown_content

    def test_cell_text_escaping(self, tmp_path):
        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "含|竖线"
            sheet["B1"] = "第一行\n第二行"

        file_path = _write_workbook(tmp_path / "escape.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        # 竖线转义防表格结构破坏；换行转 <br> 保留多行语义
        assert "含\\|竖线" in result.markdown_content
        assert "第一行<br>第二行" in result.markdown_content

    def test_embedded_image_materialized(self, tmp_path):
        pytest.importorskip("PIL")
        from PIL import Image as PilImage
        from openpyxl.drawing.image import Image as XlsxImage

        png_path = tmp_path / "stamp.png"
        PilImage.new("RGB", (8, 8), color=(255, 0, 0)).save(png_path)

        def populate(workbook):
            sheet = workbook.active
            sheet["A1"] = "证书截图"
            sheet.add_image(XlsxImage(str(png_path)), "B2")

        file_path = _write_workbook(tmp_path / "image.xlsx", populate)
        result = MarkitdownConverter().convert(file_path)

        assert len(result.images) == 1
        image_file = tmp_path / "image_images" / "sheet1_img1.png"
        assert image_file.is_file()
        assert "![工作表内嵌图片 sheet1_img1" in result.markdown_content
        assert "（约第 2 行）" in result.markdown_content
        assert "(image_images/sheet1_img1.png)" in result.markdown_content

    def test_corrupt_file_raises_friendly_error(self, tmp_path):
        file_path = tmp_path / "broken.xlsx"
        file_path.write_bytes(b"this is not a zip archive")

        with pytest.raises(MarkitdownConversionError) as exc_info:
            MarkitdownConverter().convert(file_path)
        assert "损坏" in str(exc_info.value)
        assert "加密" in str(exc_info.value)

    def test_xls_suffix_rejected(self, tmp_path):
        file_path = tmp_path / "legacy.xls"
        file_path.write_bytes(b"dummy")

        with pytest.raises(ValueError) as exc_info:
            MarkitdownConverter().convert(file_path)
        assert "暂不支持" in str(exc_info.value)
        assert "XLSX" in str(exc_info.value)

    def test_default_caps_exposed_as_constants(self):
        converter = MarkitdownConverter()
        assert converter.xlsx_max_rows == XLSX_MAX_ROWS_PER_SHEET == 300
        assert converter.xlsx_max_cols == XLSX_MAX_COLS_PER_SHEET == 64
